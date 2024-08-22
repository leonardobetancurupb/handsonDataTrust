"""
    A view in Django is essentially a Python function or class that handles incoming web requests and returns a web response. 
    This response can be anything from a simple text string to a complex HTML document, an image, a redirect, or even an error message.
"""
import openpyxl.workbook
from .serializers import UserSerializer, CategorySerializer, HolderSerializer, ConsumerSerializer, AdminSerializer, SchemaSerializer, DataSerializer, PolicySerializer, CountCollectionSerializer
from .models import CustomUser, Category, Holder, Consumer, Admin, Data, Schema, Policy, CountCollection
from rest_framework import permissions
from rest_framework import viewsets
from rest_framework.views import APIView
from rest_framework.response import Response
from django.contrib.auth import authenticate
from rest_framework_simplejwt.tokens import RefreshToken
from .decorators import admin_required, holder_required, consumer_required
from rest_framework import viewsets
from .serializers import UserSerializer
from django.contrib.auth import get_user_model
import requests
import openpyxl
import bcrypt
import pandas as pd
import json
from .operations import Operations
import datetime
import os
import bcrypt
import pandas as pd
from glob import glob
import shutil
import json
from .operations import Operations
import datetime
from django.core.files.storage import FileSystemStorage
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from io import BytesIO
from django.http import HttpResponse
from pathlib import Path 
    
@csrf_exempt
def verifyDate(request):
    headers = {'Content-Type': 'application/json'}

    if request.method=='POST':
        payloadLogin={
           "username":"admin",
           "password":"admin",
        }
        r= requests.post(f"http://127.0.0.1:8000/login/", data=payloadLogin,timeout=5)
        headersh=""
        if r.status_code<300:
            tokens=r.json()
            accesToken=tokens['access']
            token="Bearer "+accesToken
            headersh = {       
                'Content-Type': 'application/json',
                'Authorization': token
            }
        else:
            return JsonResponse({"error":"failed login"})

        consumersBody=requests.get(f"http://127.0.0.1:8000/api/consumers/", timeout=5)
        idsPolicy=[]
        consumerlst=[]

        if consumersBody.status_code!=404 and consumersBody.status_code!=500:    
            consumersData=consumersBody.json()
            dateNow=datetime.datetime.today().date()
            for consumer in consumersData:
                consumerlst.append(consumer['id'])
                for consumerData in consumer['authorization']:
                    if consumerData['finalDate']==str(dateNow):
                        idsPolicy.append(consumerData['idPolicy'])
                        consumer['authorization'].remove(consumerData)

                r= requests.request("PATCH",f"http://127.0.0.1:8000/api/consumers/{consumer['id']}/",headers=headers, data=json.dumps(consumer), timeout=5)        
            
            idsPolicy2=set(idsPolicy)
            holderBody=requests.get(f"http://127.0.0.1:8000/api/holders/", timeout=5)
            if holderBody.status_code!=404:
                holder=holderBody.json()
                for h in holder:
                    for authorization in h['authorization']:
                        if authorization['idPolicy'] in idsPolicy2:
                            h['authorization'].remove(authorization)
                    r= requests.request("PATCH",f"http://127.0.0.1:8000/api/holders/{h['id']}/",headers=headers, data=json.dumps(h))
            else:
                return JsonResponse({"error":str(holderBody.status_code)})
            for policy in idsPolicy2:
                res=requests.request("DELETE", f"http://127.0.0.1:8000/api/policy/{policy}/",headers=headersh, timeout=5)
                if 200<=res.status_code<300:                   
                    for consumer in consumerlst:    
                        path=f"./documents_consumer/{consumer}"
                        for folder in os.listdir(path):
                            completePath=os.path.join(path,folder)
                            if os.path.isdir(completePath):
                                folderSplit=folder.split("_")
                                if folderSplit[-1]==policy:
                                    shutil.rmtree(completePath)
                else:
                    return JsonResponse({"error":str(res.status_code)})
            
            dataBody=requests.get(f"http://127.0.0.1:8000/api/data/", timeout=5)
            if dataBody.status_code!=404:
                dataData=dataBody.json()
                for data in dataData:
                    if data['idPolicy'] in idsPolicy2:
                        r=requests.post(f"http://127.0.0.1:8000/deleteData/{data['id']}/", headers=headers)
                        if 200<= r.status_code <300:
                            return JsonResponse({"message":"ok"})
                        else:
                            return JsonResponse({"error":str(r.status_code)})
            else:
                return JsonResponse({"error":str(dataBody.status_code)})
        else:
            return JsonResponse({"error":str(consumersBody.status_code)})

@csrf_exempt
def saveData(request, userType, idUser):
    headers = {'Content-Type': 'application/json'}
    if request.method == 'POST':
        types=['holder']
        if userType not in types:
            return JsonResponse({'error':f'{userType} is not a valid type of user'})
        idHolderValidation= requests.get(f"http://127.0.0.1:8000/api/holders/{idUser}/", timeout=5)
        if idHolderValidation.status_code==404:
            return JsonResponse({'error':f'{userType} id doesnt exists'})
        elif 200<= idHolderValidation.status_code<300: 
            holderInfo=idHolderValidation.json()
            bodyData = request.POST.copy()
            if 'archivo' not in request.FILES:
                return JsonResponse({'error': 'file not sent'})
            file = request.FILES['archivo']
            fs = FileSystemStorage()
            schema= requests.get(f"http://127.0.0.1:8000/api/schema/{bodyData['idSchema']}/", timeout=5)
            extension=os.path.splitext(file.name)[1].lower()
            if extension !=".xlsx" and extension!=".xls" and extension!=".csv":
                return JsonResponse({"error":"not valid extension file"})
            schemaData=schema.json()
            name=schemaData['name']+"."+bodyData['format']
            path2="./documents_"+userType+"/"+idUser
            name=schemaData['name']+"."+bodyData['format']
            if os.path.exists(os.path.join(path2,name)):
                return JsonResponse({"error":"data already exist"})
            path="documents_"+userType+"/"+idUser
            if not os.path.exists(path):
                os.makedirs(path) 
            expectedHeaders= schemaData['structure'].split(" ")
            headersFile=""
            workbook=openpyxl.load_workbook(file)
            sheet=workbook.active
            for row in sheet.iter_rows(values_only=True):
                headersFile=row
                break
            if list(headersFile) != expectedHeaders:
                return JsonResponse({
                    "error":"headers file and headers schema are different",
                    "headersFile":headersFile,
                    "headersSchema":expectedHeaders
                })
            if bodyData['format']!=extension.replace(".",""):
                return JsonResponse({"error":"extension file is different than specific extension in format input"})
            filename = fs.save(os.path.join(path,name), file)
            urlFile = fs.url(filename) 
            countBody= requests.get(f"http://127.0.0.1:8000/api/count/6/", timeout=5)
            bodyCountJson=countBody.json()
            if countBody.status_code==404: 
                idCurrentCount="0"
                payload = json.dumps({
                    "id": "6",
                    "collection": "Data",
                    "count": idCurrentCount
                    })
                bodyData['id']=idCurrentCount
                bodyData['url']=urlFile
                bodyData['idHolder']=idUser
                bodyData.update(bodyData)
                serializer = DataSerializer(data=bodyData)
                r=""
                if serializer.is_valid(raise_exception=True):
                    r=requests.request("POST","http://127.0.0.1:8000/api/data/", headers=headers, data=json.dumps(bodyData), timeout=5)
                    if 200 <= r.status_code <=300:
                        holderInfo['data'].append(int(idCurrentCount))
                        Res= requests.request("PATCH",f"http://127.0.0.1:8000/api/holders/{idUser}/",headers=headers, data=json.dumps(holderInfo), timeout=5)        
                        if 200<= Res.status_code <300:
                            Operations.send_log("CREATE",{
                                "description":"Creating new data",
                                "datetime":str(datetime.datetime.now()),
                                "creationStructure":str(bodyData)
                            },idUser,"Person Collection")
                            re=requests.request("POST", "http://127.0.0.1:8000/api/count/", headers=headers, data=payload, timeout=5)
                            if 200<=re.status_code<300:
                                return JsonResponse({'message':"status code OK"})
                            else:  
                                return JsonResponse({'error':f"status code 'post' count: {re.status_code}"})
                        else:
                            return JsonResponse({'error':f"status code 'patch' holder: {Res.status_code}"})
                    else:
                        return JsonResponse({'error':f"status code 'post' data: {r.status_code}"})
                else:
                    return serializer.errors
            else:
                idCurrentCount=int(bodyCountJson['count']) 
                idCurrentCount+=1 
                bodyData['id']=idCurrentCount
                bodyData['url']=urlFile
                bodyData['idHolder']=idUser
                bodyCountJson['count']=str(idCurrentCount)
                bodyData.update(bodyData)
                serializer = DataSerializer(data=bodyData)
                r=""
                if serializer.is_valid(raise_exception=True):
                    r=requests.request("POST","http://127.0.0.1:8000/api/data/", headers=headers, data=json.dumps(bodyData), timeout=5)
                    if 200 <= r.status_code <=300:
                        holderInfo['data'].append(int(idCurrentCount))
                        Res= requests.request("PATCH",f"http://127.0.0.1:8000/api/holders/{idUser}/",headers=headers, data=json.dumps(holderInfo), timeout=5)        
                        if 200<= Res.status_code <300:
                            Operations.send_log("CREATE",{
                                "description":"Creating new data",
                                "datetime":str(datetime.datetime.now()),
                                "creationStructure":str(bodyData)
                            },idUser,"Person Collection")
                            re=requests.request("PATCH", "http://127.0.0.1:8000/api/count/6/", headers=headers, data=json.dumps(bodyCountJson), timeout=5)
                            if 200<=re.status_code<300:
                                return HttpResponse("OK")
                            else:  
                                return JsonResponse({'error':f"status code 'post' count: {re.status_code}"})
                        else:
                            return JsonResponse({'error':f"status code 'patch' holder: {Res.status_code}"})
                    else:
                        return JsonResponse({'error':f"status code 'post' data: {r.status_code}"})
                else:
                    return JsonResponse({'error':"format sent to serializer is invalid"})
        else:
            return JsonResponse({'error':f"idHolder dont exists: {idHolderValidation.status_code}"})
    elif request.method == 'GET':
        listFile=[]
        try:
            path=Path("documents_"+userType+"/"+idUser)
            for file in path.iterdir():
                if file.is_file():
                    listFile.append(file.name)
        except:
            return JsonResponse({"file List: ":listFile})
    return JsonResponse({'error': 'Method not resolved'}, status=405)

@csrf_exempt
def updateData(request,idData):
    headers = {'Content-Type': 'application/json'}
    if request.method == 'POST':
        dataBody= requests.get(f"http://127.0.0.1:8000/api/data/{idData}/", timeout=5)
        if dataBody.status_code == 404:
            return JsonResponse({"error":f"no data with id: {idData}"})
        dataData=dataBody.json()
        schemaBody=requests.get(f"http://127.0.0.1:8000/api/schema/{dataData['idSchema']}/", timeout=5)
        if 'archivo' not in request.FILES:
            return JsonResponse({'error': 'file Not Sent'})
        file = request.FILES['archivo']
        fs = FileSystemStorage()
        extension=os.path.splitext(file.name)[1].lower()
        if extension !=".xlsx" and extension!=".xls" and extension!=".csv":
            return JsonResponse({"error":"not valid extension file"})
        schemaData=schemaBody.json()
        expectedHeaders= schemaData['structure'].split(" ")
        headersFile=""
        if dataData['format']!=extension.replace(".",""):
            return JsonResponse({"error":"extension file is different than specific extension in format input"})
        workbook=openpyxl.load_workbook(file)
        sheet=workbook.active
        for row in sheet.iter_rows(values_only=True):
            headersFile=row
            break
        if list(headersFile) != expectedHeaders:
            return JsonResponse({
                "error":"headers file and headers schema are different",
                "headersFile":headersFile,
                "headersSchema":expectedHeaders
            })
        path="./"+dataData['url']
        if os.path.exists(path):
            os.remove(path)    
            filename = fs.save(path, file)
            urlFile = fs.url(filename)
            Operations.send_log("UPDATEFILE",{
                                "description":"Updating file",
                                "datetime":str(datetime.datetime.now()),
                                "UrlDataUpdated":str(path)
                            },str(dataData['idHolder']),path)
            return JsonResponse({"url": urlFile})
        else:
            return JsonResponse({"error": f"{path} not exist"})
    else:
         return JsonResponse({'error': 'Method not resolved'}, status=405)

@csrf_exempt
def deleteData(request,idData):
    headers = {'Content-Type': 'application/json'}
    if request.method == 'POST':
        dataBody= requests.get(f"http://127.0.0.1:8000/api/data/{idData}/", timeout=5)
        if dataBody.status_code == 404:
            return JsonResponse({"error":f"no data with id: {idData}"})
        dataData=dataBody.json()
        path="."+dataData['url']
        if os.path.exists(path):
            os.remove(path)
            dataBody= requests.delete(f"http://127.0.0.1:8000/api/data/{idData}/", timeout=5)
            if 200<= dataBody.status_code <300:
                return JsonResponse({"message": "Ok"})
            else:
                return JsonResponse({"error": f"{dataBody.status_code}"})
        else:
            return JsonResponse({"error": f"{path} not exist"})
    else:
         return JsonResponse({'error': 'Method not resolved'}, status=405)

@csrf_exempt
def downloadDataHolder(request,idHolder,idSchema):
    headers = {'Content-Type': 'application/json'}
    if request.method == 'POST':
        fs = FileSystemStorage()
        schemaBody= requests.get(f"http://127.0.0.1:8000/api/schema/{idSchema}/", timeout=5)
        if schemaBody==404:
            return JsonResponse({'error':f'schema with id {idSchema} not exists'})
        schemaData=schemaBody.json()
        path=f"documents_holder/{idHolder}/"
        file_paths = glob(os.path.join(fs.location, path, f"{schemaData['name']}.*"))
        if file_paths:
            full_path = file_paths[0]
            with open(full_path, 'rb') as archivo:
                response = HttpResponse(archivo.read(), content_type='application/octet-stream')
                response['Content-Disposition'] = f'attachment; filename="{os.path.basename(full_path)}"'
                return response
        else:
            return HttpResponse('file not found', status=404)

@csrf_exempt
def sign(request):
    headers = {'Content-Type': 'application/json'}
    if request.method == 'POST':
        body=json.loads(request.body)
        schemaBody= requests.get(f"http://127.0.0.1:8000/api/schema/{body['idSchema']}/", timeout=5)
        if schemaBody.status_code==404:
            return JsonResponse({"error":"there is not schema with that id"})
        policyBody= requests.get(f"http://127.0.0.1:8000/api/policy/{body['idPolicy']}/", timeout=5)
        if policyBody.status_code==404:
            return JsonResponse({"error":"there is not policy with that id"})
        schemaData=schemaBody.json()
        policyData=policyBody.json()
        schemaEncryptFields=schemaData['fieldToEncrypt']
        header=schemaData['structure'].split(" ")
        lstDataId=body['lstDataId']
        idHolderUrl="documents_holder/" 
        df=""
        workbook=openpyxl.Workbook()
        sheet=workbook.active
        sheet.append(header)
        id1=body['lstDataId'][0]
        dataBody= requests.get(f"http://127.0.0.1:8000/api/data/{id1}/", timeout=5)
        dataData=dataBody.json()
        r=""
        moneyPaid=0
        for id in lstDataId:
            dataBodyIndividual= requests.get(f"http://127.0.0.1:8000/api/data/{id}/", timeout=5)
            if dataBodyIndividual.status_code==404:
                return JsonResponse({"error":f"there is not data with id {id}"})
            dataDataIndividual = dataBodyIndividual.json()
            holderBody= requests.get(f"http://127.0.0.1:8000/api/holders/{dataDataIndividual['idHolder']}/", timeout=5)
            if holderBody.status_code==404:
                return JsonResponse({"error":"there is not holder with that id"})
            holderData=holderBody.json()
            payload={
                str(body['idConsumer']):schemaData['name'],
                "idPolicy":body['idPolicy']
            }
            holderData['authorization'].append(payload)
            holderData['money']=float(holderData['money'])+float(policyData['Value'])
            moneyPaid=float(moneyPaid)+float(policyData['Value'])
            response = requests.patch(f"http://127.0.0.1:8000/api/holders/{dataDataIndividual['idHolder']}/", headers=headers, data=json.dumps(holderData), timeout=5)
            if response.status_code != 200:
                return JsonResponse({'error': f"Failed to update holder data. Status code: {response.status_code}"})
            Operations.send_log("UPDATE CONSUMPTION",{
                                "description":"New Consumer consuming data",
                                "datetime":str(datetime.datetime.now()),
                                'schemaConsumed':str(schemaData['name']),
                                "idConsumer":str(body['idConsumer'])
                            },str(body['idConsumer']),"Holder Collection")
            idSelected=idHolderUrl+f"/{dataDataIndividual['idHolder']}/{schemaData['name']}.{dataData['format']}"
            df=pd.read_excel(idSelected)
            if len(schemaEncryptFields)!=0:
                for column in schemaEncryptFields:
                    df[column] = df[column].apply(lambda x: bcrypt.hashpw(str(x).encode('utf-8'), bcrypt.gensalt()))
            df.to_excel(idSelected, index=False)
            df1=pd.read_excel(idSelected)
            data=df1.values.tolist()
            for row in data:
                sheet.append(row)
        dateNo=datetime.datetime.now()
        path=f"documents_consumer/{body['idConsumer']}/{schemaData['name']}_{dateNo}_{body['idPolicy']}"
        if not os.path.exists(path):
                os.makedirs(path)
        destiny=path+f"/{schemaData['name']}.{dataData['format']}"
        workbook.save(destiny)
        consumerBody= requests.get(f"http://127.0.0.1:8000/api/consumers/{body['idConsumer']}/", timeout=5)
        if consumerBody==404:
            return JsonResponse({'error':'consumir with idConsumer not exists'})
        consumerData=consumerBody.json()
        payload={
            "nameSchema":schemaData['name'],
            "lstSignedData":body['lstDataId'],
            "idPolicy":policyData['id'],
            "initialDate":str(datetime.datetime.now()),
            "finalDate":policyData['estimatedTime'],
            "carpet":str(dateNo),
            "dir":path
        } 
        consumerData['authorization'].append(payload)
        consumerData['moneyPaid']=float(consumerData['moneyPaid'])+moneyPaid
        r=requests.request("PATCH",f"http://127.0.0.1:8000/api/consumers/{body['idConsumer']}/", headers=headers, data=json.dumps(consumerData), timeout=5)
        if 200<=r.status_code<300:
            Operations.send_log("UPDATE CONSUMPTION",{
                    "description":"Adding dataset to authorization",
                    "datetime":str(datetime.datetime.now()),
                    "datasetId":body['lstDataId']
                },body['idConsumer'],"Consumer Collection")
            return JsonResponse({'message':r.status_code})
        else:
            return JsonResponse({'error':r.status_code})
    else:
        return JsonResponse({'error':'method no valid'})
        
@csrf_exempt
def downloadSchema(request,idSchema):
    headers = {'Content-Type': 'application/json'}
    if request.method == 'POST':
        schemaBody= requests.get(f"http://127.0.0.1:8000/api/schema/{idSchema}/", timeout=5)
        if schemaBody.status_code==404:
            return JsonResponse({'error':f'schema with id {idSchema} not exists'})
        schemaData=schemaBody.json()
        lst = schemaData['structure'].split(" ")
        wb= openpyxl.Workbook()
        ws= wb.active
        headersXlsx=lst
        ws.append(headersXlsx)
        file_obj=BytesIO()
        wb.save(file_obj)
        file_obj.seek(0)
        name=schemaData['name']+".xlsx"
        response = HttpResponse(file_obj, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={name}'
        return response

@csrf_exempt
def downloadEncrypted(request,idConsumer,idSchema,carpet):
    headers = {'Content-Type': 'application/json'}
    if request.method == 'POST':
        fs = FileSystemStorage()
        schemaBody= requests.get(f"http://127.0.0.1:8000/api/schema/{idSchema}/", timeout=5)
        if schemaBody==404:
            return JsonResponse({'error':f'schema with id {idSchema} not exists'})
        schemaData=schemaBody.json()
        consumerBody=requests.get(f"http://127.0.0.1:8000/api/consumers/{idConsumer}/")
        consumerData=""
        if 200<=consumerBody.status_code<300:
            consumerData=consumerBody.json()
        else:
            JsonResponse({"error":str(consumerBody.status_code)})
        path=f"./"
        for authorization in consumerData['authorization']:
            if authorization['carpet']==carpet:
                path=path+authorization['dir']+"/"
        file_paths = glob(os.path.join(fs.location, path, f"{schemaData['name']}.*"))
        if file_paths:
            full_path = file_paths[0]
            with open(full_path, 'rb') as archivo:
                response = HttpResponse(archivo.read(), content_type='application/octet-stream')
                response['Content-Disposition'] = f'attachment; filename="{os.path.basename(full_path)}"'
                Operations.send_log("DOWNLOAD DATASET",{
                                "description":"Consumer downloading dataset",
                                "datetime":str(datetime.datetime.now()),
                                'schema':str(schemaData['name']),
                                "idConsumer":str(idConsumer),
                                "urlDataDownloaded":str(path)
                            },str(idConsumer),path)
                return response
        else:
            return JsonResponse({'error':'file not found'})

User = get_user_model()

class RegisterViewSet(viewsets.ModelViewSet):
    queryset = User.objects.all()
    serializer_class = UserSerializer
    def create(self, request, *args, **kwargs):
        jsonData= requests.get(f"http://127.0.0.1:8000/api/count/1/")
        data=jsonData.json()
        headers = {
                    'Content-Type': 'application/json'
                    }
        if jsonData.status_code==404:
            currentId="0"
            payload = json.dumps({
                "id": "1",
                "collection": "Person",
                "count": currentId
                })
            body=json.loads(request.body)
            body['id']=currentId
            request.data.update(body)
            serializer = UserSerializer(data=request.data)
            if serializer.is_valid(raise_exception=True):
                user = self.perform_create(serializer)
                header = self.get_success_headers(serializer.data)
                Operations.send_log("CREATE",{
                    "description":"Creating new person",
                    "datetime":str(datetime.datetime.now()),
                    "creationStructure":str(body)
                },body['id'],"Person Collection")
                requests.request("POST", "http://127.0.0.1:8000/api/count/", headers=headers, data=payload)       
                return JsonResponse({'id':currentId})
            else:
                return JsonResponse({'error':"Error in serializer"})
        else:
            currentId=int(data['count'])
            currentId+=1
            body=json.loads(request.body)
            body['id']=str(currentId)
            data['count']=str(currentId)
            request.data.update(body)
            serializer = UserSerializer(data=request.data)
            if serializer.is_valid(raise_exception=True):
                user = self.perform_create(serializer)
                header = self.get_success_headers(serializer.data)
                Operations.send_log("CREATE",{
                    "description":"Creating new person",
                    "datetime":str(datetime.datetime.now()),
                    "creationStructure":str(body)
                },body['id'],"Person Collection")
                Res= requests.request("PATCH","http://127.0.0.1:8000/api/count/1/",headers=headers, data=json.dumps(data))        
                return JsonResponse({'id':currentId})
            else:
                return JsonResponse({'error':"Error in serializer"})
    
    def update(self, request, *args, **kwargs):
        jsonData= requests.get(f"http://127.0.0.1:8000/api/registers/{kwargs['pk']}/")
        data=jsonData.json()
        body=json.loads(request.body)
        r=super().update(request, *args, **kwargs)
        if 200 <= r.status_code <= 300:
            Operations.send_log("UPDATE",{
                "description":"Updating person",
                "datetime":str(datetime.datetime.now()),
                "oldValue":data,
                "newValue":str(body)
            },"Admin","Person Collection")
            return r
        else:
            return JsonResponse({'error':f"{r.status_code}"})
    
    def destroy(self, request, *args, **kwargs):
        jsonData= requests.get(f"http://127.0.0.1:8000/api/registers/{kwargs['pk']}/")
        data=jsonData.json()
        r=super().destroy(request, *args, **kwargs)
        if 200 <= r.status_code <= 300:
            Operations.send_log("DELETE",{
                "description":"deleting person",
                "datetime":str(datetime.datetime.now()),
                "deletedValue":data
            },"Admin","Person Collection")
            return r
        else:
            return JsonResponse({'error':f"{r.status_code}"})
    
    def retrieve(self, request, *args, **kwargs):
        r=super().retrieve(request, *args, **kwargs)
        if 200 <= r.status_code <= 300:
            Operations.send_log("GET BY ID",{
                "description":"Getting Person",
                "datetime":str(datetime.datetime.now()),
                "searchedUserId":str(kwargs['pk'])
            },"Admin","Person Collection")
            return r 
        else:
            return JsonResponse({'error':f"{r.status_code}"})
    
    def list(self, request, *args, **kwargs):
        r=super().list(request, *args, **kwargs)
        if 200 <= r.status_code <= 300:    
            Operations.send_log("GET",{
                "description":"Getting People",
                "datetime":str(datetime.datetime.now())
            },"Admin","Person Collection")
            return r
        else:
            return JsonResponse({'error':f"{r.status_code}"})

    def perform_create(self, serializer):
        return serializer.save()
    
class LoginView(APIView):
    permission_classes = (permissions.AllowAny,)
    def post(self, request):
        username = request.data.get('username')
        password = request.data.get('password')
        user = authenticate(username=username, password=password)
        if user:
            refresh = RefreshToken.for_user(user)
            return Response({
                'refresh': str(refresh),
                'access': str(refresh.access_token),
                'id': str(user.id),
            })
        else:
            return Response({'error': 'Error log in'})

class HolderViewSet(viewsets.ModelViewSet):
    queryset = Holder.objects.all()
    serializer_class = HolderSerializer
    def create(self, request, *args, **kwargs):
        body=json.loads(request.body)
        validationIdPerson= requests.get(f"http://127.0.0.1:8000/api/registers/{body['idPerson']}/")
        if validationIdPerson.status_code==404:
            return JsonResponse({'error':"idPerson not found"})
        if validationIdPerson.json()['role']!="holder":
            return JsonResponse({'error':"idPerson with role different to holder"})
        jsonData= requests.get(f"http://127.0.0.1:8000/api/count/2/")
        data=jsonData.json()
        headers = {
                    'Content-Type': 'application/json'
                    }
        if jsonData.status_code==404:
            currentId="0"
            payload = json.dumps({
                "id": "2",
                "collection": "Holder",
                "count": currentId
                })
            body['id']=currentId
            body['data']=[]
            body['authorization']=[]
            body['money']=0     
            request.data.update(body)
            serializer = HolderSerializer(data=request.data)
            r=""
            if serializer.is_valid(raise_exception=True):
                r=super().create(request, *args, **kwargs)
                if 200<=r.status_code<300:
                    Operations.send_log("CREATE",{
                        "description":"Creating new holder",
                        "datetime":str(datetime.datetime.now()),
                        "creationStructure":str(body)
                    },body['id'],"Holder Collection")
                    requests.request("POST", "http://127.0.0.1:8000/api/count/", headers=headers, data=payload)    
                    return r   
                else:   
                    return JsonResponse({'error':f"holder not created; status code: {r.status_code}"})
            else:
                return JsonResponse({'error':"Serializer format not valid"})
        else:
            currentId=int(data['count'])
            currentId+=1
            body['id']=str(currentId)
            body['data']=[]
            body['authorization']=[]
            body['money']=0    
            data['count']=str(currentId)
            request.data.update(body)
            serializer = HolderSerializer(data=request.data)
            r=""
            if serializer.is_valid(raise_exception=True):
                r=super().create(request, *args, **kwargs)
                if 200<=r.status_code<300:
                    Operations.send_log("CREATE",{
                        "description":"Creating new holder",
                        "datetime":str(datetime.datetime.now()),
                        "creationStructure":str(body)
                    },body['id'],"Holder Collection")
                    Res= requests.request("PATCH","http://127.0.0.1:8000/api/count/2/",headers=headers, data=json.dumps(data))        
                    return r
                else:
                    return JsonResponse({'error':f"Holder not created; status code: {r.status_code}"}) 
            else:
                return JsonResponse({'error':"Serializer format not valid"})
    
    def update(self, request, *args, **kwargs):
        jsonData= requests.get(f"http://127.0.0.1:8000/api/holders/{kwargs['pk']}/")
        data=jsonData.json()
        body=json.loads(request.body)
        r=super().update(request, *args, **kwargs)
        if 200 <= r.status_code <= 300:
            print(Operations.send_log("UPDATE",{
                "description":"Updating holder",
                "datetime":str(datetime.datetime.now()),
                "oldValue":data,
                "newValue":str(body)
            },data['id'],"Holder Collection"))
            return r
        else:
            return r

    def destroy(self, request, *args, **kwargs):
        jsonData= requests.get(f"http://127.0.0.1:8000/api/holders/{kwargs['pk']}/")
        data=jsonData.json()
        r=super().destroy(request, *args, **kwargs)
        if 200 <= r.status_code <= 300:
            print(Operations.send_log("DELETE",{
                "description":"deleting holder",
                "datetime":str(datetime.datetime.now()),
                "deletedValue":data
            },data['id'],"Holder Collection"))
            return r
        else:
            return r 

    def retrieve(self, request, *args, **kwargs):
        r=super().retrieve(request, *args, **kwargs)
        if 200 <= r.status_code <= 300:
            print(Operations.send_log("GET BY ID",{
                "description":"Getting Holder",
                "datetime":str(datetime.datetime.now()),
                "searchedHolderId":str(kwargs['pk'])
            },"Admin","Holder Collection"))
            return r 
        else:
            return r

    def list(self, request, *args, **kwargs):
        r=super().list(request, *args, **kwargs)
        if 200 <= r.status_code <= 300:    
            print(Operations.send_log("GET",{
                "description":"Getting holders",
                "datetime":str(datetime.datetime.now())
            },"Admin","Holder Collection"))
            return r
        else:
            return r
    
class ConsumerViewSet(viewsets.ModelViewSet):
    queryset = Consumer.objects.all()
    serializer_class = ConsumerSerializer
    def create(self, request, *args, **kwargs):
        body=json.loads(request.body)
        validationIdPerson= requests.get(f"http://127.0.0.1:8000/api/registers/{body['idPerson']}/")
        if validationIdPerson.status_code==404:
            return JsonResponse({'error':"idPerson not found"})
        if validationIdPerson.json()['role']!="consumer":
            return JsonResponse({'error':"idPerson with role different to consumer"})
        jsonData= requests.get(f"http://127.0.0.1:8000/api/count/3/")
        data=jsonData.json()
        headers = {
                    'Content-Type': 'application/json'
                    }
        if jsonData.status_code==404:
            currentId="0"
            payload = json.dumps({
                "id": "3",
                "collection": "Consumer",
                "count": currentId
                })
            body['id']=currentId
            body['authorization']=[]  
            body['moneyPaid']=0
            request.data.update(body)
            serializer = ConsumerSerializer(data=request.data)
            r=""
            if serializer.is_valid(raise_exception=True):
                r=super().create(request, *args, **kwargs)
                if 200<=r.status_code<300:
                    Operations.send_log("CREATE",{
                        "description":"Creating new consumer",
                        "datetime":str(datetime.datetime.now()),
                        "creationStructure":str(body)
                    },body['id'],"Consumer Collection")
                    requests.request("POST", "http://127.0.0.1:8000/api/count/", headers=headers, data=payload)    
                    return r
                else:
                    return JsonResponse({'error':f"Consumer not created; status code: {r.status_code}"})
            else:
                return JsonResponse({'error':"Serializer format not valid"})
        else:
            currentId=int(data['count'])
            currentId+=1
            body['id']=str(currentId)
            body['authorization']=[]
            body['moneyPaid']=0
            data['count']=str(currentId)
            request.data.update(body)
            serializer = ConsumerSerializer(data=request.data)
            r=""
            if serializer.is_valid(raise_exception=True):
                r=super().create(request, *args, **kwargs)
                if 200<=r.status_code<300:
                    Operations.send_log("CREATE",{
                        "description":"Creating new consumer",
                        "datetime":str(datetime.datetime.now()),
                        "creationStructure":str(body)
                    },body['id'],"Consumer Collection")
                    Res= requests.request("PATCH","http://127.0.0.1:8000/api/count/3/",headers=headers, data=json.dumps(data))        
                    return r
                else:
                    return JsonResponse({'error':f"Consumer not created; status code: {r.status_code}"})
            else:
                return JsonResponse({'error':"Serializer format not valid"})
            
    def update(self, request, *args, **kwargs):
        jsonData= requests.get(f"http://127.0.0.1:8000/api/consumers/{kwargs['pk']}/")
        data=jsonData.json()
        body=json.loads(request.body)
        r=super().update(request, *args, **kwargs)
        if 200 <= r.status_code <= 300:
            Operations.send_log("UPDATE",{
                "description":"Updating consumer",
                "datetime":str(datetime.datetime.now()),
                "oldValue":data,
                "newValue":str(body)
            },data['id'],"Consumer Collection")
            return r
        else:
            return JsonResponse({'error':f"Consumer not updated; status code: {r.status_code}"})
    
    def destroy(self, request, *args, **kwargs):
        jsonData= requests.get(f"http://127.0.0.1:8000/api/consumers/{kwargs['pk']}/")
        data=jsonData.json()
        r=super().destroy(request, *args, **kwargs)
        if 200 <= r.status_code <= 300:
            Operations.send_log("DELETE",{
                "description":"deleting consumer",
                "datetime":str(datetime.datetime.now()),
                "deletedValue":data
            },data['id'],"Consumer Collection")
            return r
        else:
            return JsonResponse({'error':f"Consumer not destroy; status_code: {r.status_code}"})

    def retrieve(self, request, *args, **kwargs):
        r=super().retrieve(request, *args, **kwargs)
        if 200 <= r.status_code <= 300:
            Operations.send_log("GET BY ID",{
                "description":"Getting Consumer",
                "datetime":str(datetime.datetime.now()),
                "searchedConsumerId":str(kwargs['pk'])
            },"Admin","Consumer Collection")
            return r 
        else:
            return JsonResponse({'error':f"not returned consumer; status code: {r.status_code}"})
    
    def list(self, request, *args, **kwargs):
        r=super().list(request, *args, **kwargs)
        if 200 <= r.status_code <= 300:    
            Operations.send_log("GET",{
                "description":"Getting consumers",
                "datetime":str(datetime.datetime.now())
            },"Admin","Consumer Collection")
            return r
        else:
            return JsonResponse({'error':f"not returned data; status_code {r.status_code}"})

class AdminViewSet(viewsets.ModelViewSet):
    queryset = Admin.objects.all()
    serializer_class = AdminSerializer
    def create(self, request, *args, **kwargs):
        body=json.loads(request.body)
        validationIdPerson= requests.get(f"http://127.0.0.1:8000/api/registers/{body['idPerson']}/")
        if validationIdPerson.status_code==404:
            return JsonResponse({'error':"idPerson not found"})
        if validationIdPerson.json()['role']!="admin":
            return JsonResponse({'error':"idPerson with role different to admin"})
        jsonData= requests.get(f"http://127.0.0.1:8000/api/count/4/")
        data=jsonData.json()
        headers = {
                    'Content-Type': 'application/json'
                    }
        if jsonData.status_code==404:
            currentId="0"
            payload = json.dumps({
                "id": "4",
                "collection": "Admin",
                "count": currentId
                })
            body['id']=currentId 
            request.data.update(body)
            serializer = AdminSerializer(data=request.data)
            r=""
            if serializer.is_valid(raise_exception=True):
                r=super().create(request, *args, **kwargs)
                if 200<=r.status_code<300:
                    Operations.send_log("CREATE",{
                        "description":"Creating new Admin",
                        "datetime":str(datetime.datetime.now()),
                        "creationStructure":str(body)
                    },body['id'],"Admin Collection")
                    requests.request("POST", "http://127.0.0.1:8000/api/count/", headers=headers, data=payload)    
                    return r   
                else:
                    return JsonResponse({'error':f"Admin not created; status code: {r.status_code}"})
            else:
                return JsonResponse({'error':"Serializer format not valid"})
        else:
            currentId=int(data['count'])
            currentId+=1
            body['id']=str(currentId)
            data['count']=str(currentId)
            request.data.update(body)
            serializer = AdminSerializer(data=request.data)
            r=""
            if serializer.is_valid(raise_exception=True):
                r=super().create(request, *args, **kwargs)
                if 200<=r.status_code<300:
                    Operations.send_log("CREATE",{
                        "description":"Creating new admin",
                        "datetime":str(datetime.datetime.now()),
                        "creationStructure":str(body)
                    },body['id'],"Admin Collection")
                    Res= requests.request("PATCH","http://127.0.0.1:8000/api/count/4/",headers=headers, data=json.dumps(data))        
                    return r
                else:
                    return JsonResponse({'error':f"Admin not created; status code: {r.status_code}"}) 
            else:
                return JsonResponse({'error':"Serializer format not valid"})
            
    def update(self, request, *args, **kwargs):
        jsonData= requests.get(f"http://127.0.0.1:8000/api/admin/{kwargs['pk']}/")
        data=jsonData.json()
        body=json.loads(request.body)
        r=super().update(request, *args, **kwargs)
        if 200 <= r.status_code <= 300:
            Operations.send_log("UPDATE",{
                "description":"Updating admin",
                "datetime":str(datetime.datetime.now()),
                "oldValue":data,
                "newValue":str(body)
            },data['id'],"Admin Collection")
            return r
        else:
            return JsonResponse({'error':f"update admin failed; status code: {r.status_code}"})

    def destroy(self, request, *args, **kwargs):
        jsonData= requests.get(f"http://127.0.0.1:8000/api/admin/{kwargs['pk']}/")
        data=jsonData.json()
        r=super().destroy(request, *args, **kwargs)
        if 200 <= r.status_code <= 300:
            Operations.send_log("DELETE",{
                "description":"deleting admin",
                "datetime":str(datetime.datetime.now()),
                "deletedValue":data
            },data['id'],"Admin Collection")
            return r
        else:
            return JsonResponse({'error':f"destroy admin failed; status_code {r.status_code}"}) 

    def retrieve(self, request, *args, **kwargs):
        r=super().retrieve(request, *args, **kwargs)
        if 200 <= r.status_code <= 300:
            Operations.send_log("GET BY ID",{
                "description":"Getting Admin",
                "datetime":str(datetime.datetime.now()),
                "searchedAdminId":str(kwargs['pk'])
            },"Admin","Admin Collection")
            return r 
        else:
            return JsonResponse({'error':f"retrieve admin failed, status code: {r.status_code}"})
    
    def list(self, request, *args, **kwargs):
        r=super().list(request, *args, **kwargs)
        if 200 <= r.status_code <= 300:    
            Operations.send_log("GET",{
                "description":"Getting admin",
                "datetime":str(datetime.datetime.now())
            },"Admin","Admin Collection")
            return r
        else:
            return JsonResponse({'error':f"return admins failed; status code: {r.status_code}"})

class PolicyViewSet(viewsets.ModelViewSet):
    queryset = Policy.objects.all()
    serializer_class = PolicySerializer
    @admin_required
    def create(self, request, *args, **kwargs):
        jsonData= requests.get(f"http://127.0.0.1:8000/api/count/5/")
        data=jsonData.json()
        headers = {
                    'Content-Type': 'application/json'
                    }
        if jsonData.status_code==404:
            currentId="0"
            payload = json.dumps({
                "id": "5",
                "collection": "Policy",
                "count": currentId
                })
            body=json.loads(request.body)
            body['id']=currentId
            request.data.update(body)
            r=super().create(request, *args, **kwargs)
            if 200 <= r.status_code <= 202:
                Operations.send_log("CREATE",{
                    "description":"Creating new policy",
                    "datetime":str(datetime.datetime.now()),
                    "creationStructure":str(body)
                },"Admin","policy Collection")
                
                requests.request("POST", "http://127.0.0.1:8000/api/count/", headers=headers, data=payload)       
                return r
            else:
                return JsonResponse({'error':f"failed creating policy; status code: {r.status_code}"})
        else:      
            currentId=int(data['count'])
            currentId+=1
            body=json.loads(request.body)
            body['id']=str(currentId)
            data['count']=str(currentId)
            request.data.update(body)
            r=super().create(request, *args, **kwargs)
            
            if 200 <= r.status_code <= 202:
                Operations.send_log("CREATE",{
                    "description":"Creating new policy",
                    "datetime":str(datetime.datetime.now()),
                    "creationStructure":str(body)
                },"Admin","Policy Collection")
                Response= requests.request("PATCH","http://127.0.0.1:8000/api/count/5/",headers=headers, data=json.dumps(data))        
                return r
            else:
                return JsonResponse({'error':f"failed creating policy; status code: {r.status_code}"})

    @admin_required
    def update(self, request, *args, **kwargs):
        jsonData= requests.get(f"http://127.0.0.1:8000/api/policy/{kwargs['pk']}/")
        data=jsonData.json()
        body=json.loads(request.body)
        r=super().update(request, *args, **kwargs)
        if 200 <= r.status_code <= 300:
            Operations.send_log("UPDATE",{
                "description":"Updating policy",
                "datetime":str(datetime.datetime.now()),
                "oldValue":data,
                "newValue":str(body)
            },"Admin","Policy Collection")
            return r
        else:
            return JsonResponse({'error':f"failed updating policy; status code: {r.status_code}"})

    @admin_required
    def destroy(self, request, *args, **kwargs):
        jsonData= requests.get(f"http://127.0.0.1:8000/api/policy/{kwargs['pk']}/")
        data=jsonData.json()
        r=super().destroy(request, *args, **kwargs)
        if 200 <= r.status_code <= 300:
            Operations.send_log("DELETE",{
                "description":"deleting policy",
                "datetime":str(datetime.datetime.now()),
                "deletedValue":data
            },"Admin","Policy Collection")
            return r
        else:
            return JsonResponse({'error':f"failed destroying policy; status code: {r.status_code}"})

    def retrieve(self, request, *args, **kwargs):
        r=super().retrieve(request, *args, **kwargs)
        if 200 <= r.status_code <= 300:
            Operations.send_log("GET BY ID",{
                "description":"Getting policy",
                "datetime":str(datetime.datetime.now()),
                "searchedPolicyId":str(kwargs['pk'])
            },"Admin","Policy Collection")
            return r 
        else:
            return JsonResponse({'error':f"failed retrieving policy; status code: {r.status_code}"})

    def list(self, request, *args, **kwargs):
        r=super().list(request, *args, **kwargs)
        if 200 <= r.status_code <= 300:    
            Operations.send_log("GET",{
                "description":"Getting policies",
                "datetime":str(datetime.datetime.now())
            },"Admin","Policy Collection")
            return r
        else:
            return JsonResponse({'error':f"failed list policy; status code: {r.status_code}"})

class DataViewSet(viewsets.ModelViewSet):
    queryset = Data.objects.all()
    serializer_class = DataSerializer
    
    def create(self, request, *args, **kwargs):
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        return JsonResponse({"error":"no valid method"})
    
    def destroy(self, request, *args, **kwargs):
        headers = {
                    'Content-Type': 'application/json'
                    }
        jsonData= requests.get(f"http://127.0.0.1:8000/api/data/{kwargs['pk']}/")
        data=jsonData.json()
        jsonDataHolder= requests.get(f"http://127.0.0.1:8000/api/holders/{data['idHolder']}/")
        dataHolder=jsonDataHolder.json()
        dataHolder['data'].remove(int(data['id']))
        res= requests.request("PATCH",f"http://127.0.0.1:8000/api/holders/{data['idHolder']}/",headers=headers, data=json.dumps(dataHolder))        
        if 200<=res.status_code<300:
            r=super().destroy(request, *args, **kwargs)
            if 200<=r.status_code<300:
                Operations.send_log("DELETE",{
                                "description":"Deleting data",
                                "datetime":str(datetime.datetime.now()),
                                "deletedValue":data
                            },str(data['idHolder']),"Data Collection")
                return r
            else:
                return JsonResponse({'error':f"failed destroying Data; status code: {r.status_code}"})
        else:
            return JsonResponse({'error':f"failed updating holders collection; status code: {res.status_code}"})

    def retrieve(self, request, *args, **kwargs):
        r=super().retrieve(request, *args, **kwargs)
        if 200<=r.status_code<300:
            return r
        else:
            return JsonResponse({'error':f"failed retrieving Data; status code: {r.status_code}"})      
    
    def list(self, request, *args, **kwargs):
        r=super().list(request, *args, **kwargs)
        if 200<=r.status_code<300:
            Operations.send_log("GET",{
                "description":"Getting data list",
                "datetime":str(datetime.datetime.now())
            },"user","Data Collection")
            return r
        else:
            return JsonResponse({'error':f"failed list data; status code: {r.status_code}"})

class CategoryViewSet(viewsets.ModelViewSet):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer
    @admin_required
    def create(self, request, *args, **kwargs):
        jsonData= requests.get(f"http://127.0.0.1:8000/api/count/7/")
        data=jsonData.json()
        headers = {
                    'Content-Type': 'application/json'
                    }
        if jsonData.status_code==404:
            currentId="0"
            payload = json.dumps({
                "id": "7",
                "collection": "Category",
                "count": currentId
                })
            body=json.loads(request.body)
            body['id']=currentId
            request.data.update(body)
            r=super().create(request, *args, **kwargs)
            if 200 <= r.status_code <= 202:
                Operations.send_log("CREATE",{
                    "description":"Creating new Category",
                    "datetime":str(datetime.datetime.now()),
                    "creationStructure":str(body)
                },"Admin","Category Collection")
                requests.request("POST", "http://127.0.0.1:8000/api/count/", headers=headers, data=payload)       
                return r
            else:
                return JsonResponse({'error':f"failed creating category; status code: {r.status_code}"})
        else:      
            currentId=int(data['count'])
            currentId+=1
            body=json.loads(request.body)
            body['id']=str(currentId)
            data['count']=str(currentId)
            request.data.update(body)
            r=super().create(request, *args, **kwargs)
            if 200 <= r.status_code <= 202: 
                Operations.send_log("CREATE",{
                    "description":"Creating new Category",
                    "datetime":str(datetime.datetime.now()),
                    "creationStructure":str(body)
                },"Admin","Category Collection")
                Response= requests.request("PATCH","http://127.0.0.1:8000/api/count/7/",headers=headers, data=json.dumps(data))        
                return r
            else:
                return JsonResponse({'error':f"failed creating category; status code: {r.status_code}"})

    @admin_required
    def update(self, request, *args, **kwargs):
        jsonData= requests.get(f"http://127.0.0.1:8000/api/category/{kwargs['pk']}/")
        data=jsonData.json()
        body=json.loads(request.body)
        r=super().update(request, *args, **kwargs)
        if 200 <= r.status_code <= 300:
            Operations.send_log("UPDATE",{
                "description":"Updating category",
                "datetime":str(datetime.datetime.now()),
                "oldValue":data,
                "newValue":str(body)
            },"Admin","Category Collection")
            return r
        else:
            return JsonResponse({'error':f"failed updating category; status code: {r.status_code}"})

    @admin_required
    def destroy(self, request, *args, **kwargs):
        jsonData= requests.get(f"http://127.0.0.1:8000/api/category/{kwargs['pk']}/")
        data=jsonData.json()
        r=super().destroy(request, *args, **kwargs)
        if 200 <= r.status_code <= 300:
            Operations.send_log("DELETE",{
                "description":"deleting category",
                "datetime":str(datetime.datetime.now()),
                "deletedValue":data
            },"Admin","Category Collection")
            return r
        else:
            return JsonResponse({'error':f"failed destroying category; status code: {r.status_code}"})

    def retrieve(self, request, *args, **kwargs):
        r=super().retrieve(request, *args, **kwargs)
        if 200 <= r.status_code <= 300:
            Operations.send_log("GET BY ID",{
                "description":"Getting category",
                "datetime":str(datetime.datetime.now()),
                "searchedCategoryId":str(kwargs['pk'])
            },"Admin","Category Collection")
            return r 
        else:
            return JsonResponse({'error':f"failed retrieving category; status code: {r.status_code}"})
            
    def list(self, request, *args, **kwargs):
        r=super().list(request, *args, **kwargs)
        if 200 <= r.status_code <= 300:    
            Operations.send_log("GET",{
                "description":"Getting categories",
                "datetime":str(datetime.datetime.now())
            },"Admin","Category Collection")
            return r
        else:
            return JsonResponse({'error':f"failed list category; status code: {r.status_code}"})

class SchemaViewSet(viewsets.ModelViewSet):
    queryset = Schema.objects.all()
    serializer_class = SchemaSerializer
    @admin_required
    def create(self, request, *args, **kwargs):
        jsonData= requests.get(f"http://127.0.0.1:8000/api/count/8/")
        data=jsonData.json()
        headers = {
                    'Content-Type': 'application/json'
                    }
        if jsonData.status_code==404:
            currentId="0"
            payload = json.dumps({
                "id": "8",
                "collection": "Schema",
                "count": currentId
                })
            body=json.loads(request.body)
            body['id']=currentId
            request.data.update(body)
            r=super().create(request, *args, **kwargs)
            if 200 <= r.status_code <= 202:
                Operations.send_log("CREATE",{
                    "description":"Creating new schema",
                    "datetime":str(datetime.datetime.now()),
                    "creationStructure":{
                        "id":str(body['id']),
                        "name": str(body['name']),
                        "structure": str(body['structure']),
                        "description":str(body['description'])
                    }
                },"Admin","Schema Collection")
                requests.request("POST", "http://127.0.0.1:8000/api/count/", headers=headers, data=payload)       
                return r
            else:
                return JsonResponse({'error':f"failed creating schema; status code: {r.status_code}"})
        else:      
            currentId=int(data['count'])
            currentId+=1
            body=json.loads(request.body)
            body['id']=str(currentId)
            data['count']=str(currentId)
            request.data.update(body)
            r=super().create(request, *args, **kwargs)
            if 200 <= r.status_code <= 202:
                Operations.send_log("CREATE",{
                    "description":"Creating new schema",
                    "datetime":str(datetime.datetime.now()),
                    "creationStructure":{
                        "id":str(body['id']),
                        "name": str(body['name']),
                        "structure": str(body['structure']),
                        "description":str(body['description'])
                    }
                },"Admin","Schema Collection")
                Response= requests.request("PATCH","http://127.0.0.1:8000/api/count/8/",headers=headers, data=json.dumps(data))        
                return r
            else:
                return JsonResponse({'error':f"failed creating schema; status code: {r.status_code}"})
            
    @admin_required
    def update(self, request, *args, **kwargs):
        jsonData= requests.get(f"http://127.0.0.1:8000/api/schema/{kwargs['pk']}/")
        data=jsonData.json()
        body=json.loads(request.body)
        r=super().update(request, *args, **kwargs)
        if 200 <= r.status_code <= 202:
            Operations.send_log("UPDATE",{
                "description":"Updating schema",
                "datetime":str(datetime.datetime.now()),
                "oldValue":data,
                "newValue":str(body)
            },"Admin","Schema Collection")
            return r
        else:
            return JsonResponse({'error':f"failed updating schema; status code: {r.status_code}"})

    @admin_required    
    def destroy(self, request, *args, **kwargs):
        jsonData= requests.get(f"http://127.0.0.1:8000/api/schema/{kwargs['pk']}/")
        data=jsonData.json()
        r=super().destroy(request, *args, **kwargs)
        if 200 <= r.status_code < 300:
            Operations.send_log("DELETE",{
                "description":"deleting schema",
                "datetime":str(datetime.datetime.now()),
                "deletedValue":data
            },"Admin","Schema Collection")
            return r
        else:
            return JsonResponse({'error':f"failed destroying schema; status code: {r.status_code}"})

    def retrieve(self, request, *args, **kwargs):
        r=super().retrieve(request, *args, **kwargs)
        if 200 <= r.status_code <= 202:
            Operations.send_log("GET BY ID",{
                "description":"Getting schema",
                "datetime":str(datetime.datetime.now()),
                "searchedUserId":str(kwargs['pk'])
            },"Admin","Schema Collection")
            return r 
        else:
            return JsonResponse({'error':f"failed retrieving schema; status code: {r.status_code}"})

    def list(self, request, *args, **kwargs):
        r=super().list(request, *args, **kwargs)
        if 200 <= r.status_code <= 202:    
            Operations.send_log("GET",{
                "description":"Getting schemas",
                "datetime":str(datetime.datetime.now())
            },"Admin","Schema Collection")
            return r
        else:
            return JsonResponse({'error':f"failed list schema; status code: {r.status_code}"})
            
class CountCollectionViewSet(viewsets.ModelViewSet):
    queryset = CountCollection.objects.all()
    serializer_class = CountCollectionSerializer
    def create(self, request, *args, **kwargs):
        return super().create(request, *args, **kwargs)
    
    def update(self, request, *args, **kwargs):
        return super().update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        return super().destroy(request, *args, **kwargs)

    def retrieve(self, request, *args, **kwargs):
        return super().retrieve(request, *args, **kwargs)
    
    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)

